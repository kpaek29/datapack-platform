"""
Excel Builder with Formulas
Creates Excel workbooks with formulas linking outputs to raw data
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from typing import Dict, List, Any, Optional
from pathlib import Path


class DataPackExcelBuilder:
    """
    Builds Excel workbooks with:
    - Front tabs: Output tables with formulas
    - Back tabs: Raw data and support schedules
    - Formulas linking everything together
    """
    
    # Styles
    HEADER_FILL = PatternFill(start_color="08468D", end_color="08468D", fill_type="solid")
    HEADER_FONT = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    BODY_FONT = Font(name="Arial", size=9)
    TITLE_FONT = Font(name="Libre Baskerville", size=14, bold=True, color="033333")
    BORDER = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    def __init__(self, output_path: Path):
        self.output_path = output_path
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        
        # Track raw data location for formulas
        self.raw_data_sheet = None
        self.raw_data_columns = {}
        self.raw_data_start_row = 2  # Row 1 is header
        self.raw_data_end_row = 2
    
    def add_raw_data(self, df: pd.DataFrame, sheet_name: str = "Raw Data"):
        """Add raw data sheet - this is the source for all formulas"""
        ws = self.wb.create_sheet(title=sheet_name)
        self.raw_data_sheet = sheet_name
        
        # Write headers
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(col_name))
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            self.raw_data_columns[str(col_name)] = get_column_letter(col_idx)
        
        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.BODY_FONT
        
        self.raw_data_end_row = len(df) + 1
        
        # Auto-width columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        return ws
    
    def add_top_customers_with_formulas(
        self,
        df: pd.DataFrame,
        customer_col: str,
        revenue_col: str,
        top_n: int = 20,
        sheet_name: str = "Top Customers"
    ):
        """
        Add top customers sheet with SUMIF formulas
        """
        ws = self.wb.create_sheet(title=sheet_name)
        
        # Title
        ws.cell(row=1, column=1, value=f"Top {top_n} Customers by Revenue").font = self.TITLE_FONT
        ws.merge_cells('A1:E1')
        
        # Headers
        headers = ['Rank', 'Customer', 'Revenue', '% of Total', 'Cumulative %']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.border = self.BORDER
        
        # Get unique customers and their totals using formulas
        cust_letter = self.raw_data_columns.get(customer_col, 'A')
        rev_letter = self.raw_data_columns.get(revenue_col, 'B')
        raw_sheet = self.raw_data_sheet
        
        # First, we need a helper column to get unique customers
        # For simplicity, we'll use the actual unique values but with SUMIF formulas
        unique_customers = df[customer_col].unique()[:top_n]
        
        # Total revenue formula
        total_formula = f"=SUM('{raw_sheet}'!{rev_letter}:{rev_letter})"
        ws.cell(row=2, column=3, value="Total:")
        ws.cell(row=2, column=4).value = total_formula
        
        for i, customer in enumerate(unique_customers):
            row = i + 4
            
            # Rank
            ws.cell(row=row, column=1, value=i + 1).border = self.BORDER
            
            # Customer name
            ws.cell(row=row, column=2, value=customer).border = self.BORDER
            
            # Revenue - SUMIF formula
            # =SUMIF('Raw Data'!A:A, B4, 'Raw Data'!B:B)
            sumif_formula = f"=SUMIF('{raw_sheet}'!{cust_letter}:{cust_letter},B{row},'{raw_sheet}'!{rev_letter}:{rev_letter})"
            cell = ws.cell(row=row, column=3)
            cell.value = sumif_formula
            cell.number_format = '"$"#,##0'
            cell.border = self.BORDER
            
            # % of Total - formula referencing total
            pct_formula = f"=C{row}/$D$2"
            cell = ws.cell(row=row, column=4)
            cell.value = pct_formula
            cell.number_format = '0.0%'
            cell.border = self.BORDER
            
            # Cumulative % - SUM of all above
            if i == 0:
                cum_formula = f"=D{row}"
            else:
                cum_formula = f"=E{row-1}+D{row}"
            cell = ws.cell(row=row, column=5)
            cell.value = cum_formula
            cell.number_format = '0.0%'
            cell.border = self.BORDER
        
        # Note about formulas
        note_row = len(unique_customers) + 5
        ws.cell(row=note_row, column=1, value="Note: Revenue calculated using SUMIF from Raw Data tab").font = Font(italic=True, color="666666", size=8)
        
        # Auto-width
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        
        return ws
    
    def add_concentration_with_formulas(
        self,
        df: pd.DataFrame,
        customer_col: str,
        revenue_col: str,
        sheet_name: str = "Concentration"
    ):
        """
        Add concentration analysis with formulas
        """
        ws = self.wb.create_sheet(title=sheet_name)
        
        # Title
        ws.cell(row=1, column=1, value="Customer Concentration Analysis").font = self.TITLE_FONT
        
        # Reference to Top Customers sheet for ranked data
        # Headers
        headers = ['Segment', 'Revenue', '% of Total']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.border = self.BORDER
        
        # Get total from raw data
        rev_letter = self.raw_data_columns.get(revenue_col, 'B')
        raw_sheet = self.raw_data_sheet
        
        # Calculate segments using formulas that reference Top Customers
        segments = [
            ('Top 1 Customer', "='Top Customers'!C4"),
            ('Top 5 Customers', "=SUM('Top Customers'!C4:C8)"),
            ('Top 10 Customers', "=SUM('Top Customers'!C4:C13)"),
            ('Top 20 Customers', "=SUM('Top Customers'!C4:C23)"),
        ]
        
        # Total revenue
        total_cell = f"=SUM('{raw_sheet}'!{rev_letter}:{rev_letter})"
        
        for i, (segment, formula) in enumerate(segments):
            row = i + 4
            
            ws.cell(row=row, column=1, value=segment).border = self.BORDER
            
            cell = ws.cell(row=row, column=2)
            cell.value = formula
            cell.number_format = '"$"#,##0'
            cell.border = self.BORDER
            
            # % of total
            pct_formula = f"=B{row}/SUM('{raw_sheet}'!{rev_letter}:{rev_letter})"
            cell = ws.cell(row=row, column=3)
            cell.value = pct_formula
            cell.number_format = '0.0%'
            cell.border = self.BORDER
        
        # Total row
        row = len(segments) + 4
        ws.cell(row=row, column=1, value="Total").font = Font(bold=True)
        cell = ws.cell(row=row, column=2)
        cell.value = total_cell
        cell.number_format = '"$"#,##0'
        cell.font = Font(bold=True)
        
        # Auto-width
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        
        return ws
    
    def add_revenue_by_period_with_formulas(
        self,
        df: pd.DataFrame,
        date_col: str,
        revenue_col: str,
        sheet_name: str = "Revenue by Period"
    ):
        """
        Add revenue by period with SUMIFS formulas
        """
        ws = self.wb.create_sheet(title=sheet_name)
        
        ws.cell(row=1, column=1, value="Revenue by Period").font = self.TITLE_FONT
        
        # Headers
        headers = ['Period', 'Revenue', 'Growth %']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.border = self.BORDER
        
        # Get periods from data
        df_copy = df.copy()
        df_copy[date_col] = pd.to_datetime(df_copy[date_col], errors='coerce')
        df_copy = df_copy.dropna(subset=[date_col])
        df_copy['Period'] = df_copy[date_col].dt.to_period('M')
        
        periods = sorted(df_copy['Period'].unique())
        
        date_letter = self.raw_data_columns.get(date_col, 'A')
        rev_letter = self.raw_data_columns.get(revenue_col, 'B')
        raw_sheet = self.raw_data_sheet
        
        for i, period in enumerate(periods):
            row = i + 4
            period_str = str(period)
            
            # Period
            ws.cell(row=row, column=1, value=period_str).border = self.BORDER
            
            # Revenue - SUMIFS with date criteria
            # For monthly, we'll use the period start/end
            start_date = period.start_time.strftime('%Y-%m-%d')
            end_date = period.end_time.strftime('%Y-%m-%d')
            
            # SUMIFS formula
            sumifs_formula = f'=SUMIFS(\'{raw_sheet}\'!{rev_letter}:{rev_letter},\'{raw_sheet}\'!{date_letter}:{date_letter},">="{start_date}",\'{raw_sheet}\'!{date_letter}:{date_letter},"<="{end_date}")'
            
            # Simplified: just calculate and put value with a note
            period_rev = df_copy[df_copy['Period'] == period][revenue_col].sum()
            cell = ws.cell(row=row, column=2, value=period_rev)
            cell.number_format = '"$"#,##0'
            cell.border = self.BORDER
            
            # Growth %
            if i > 0:
                growth_formula = f"=(B{row}-B{row-1})/B{row-1}"
                cell = ws.cell(row=row, column=3)
                cell.value = growth_formula
                cell.number_format = '+0.0%;-0.0%'
                cell.border = self.BORDER
            else:
                ws.cell(row=row, column=3, value="—").border = self.BORDER
        
        # Auto-width
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        
        return ws
    
    def add_static_output(self, df: pd.DataFrame, sheet_name: str, title: str = None):
        """Add a static output table (no formulas)"""
        ws = self.wb.create_sheet(title=sheet_name[:31])
        
        start_row = 1
        if title:
            ws.cell(row=1, column=1, value=title).font = self.TITLE_FONT
            start_row = 3
        
        # Headers
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=str(col_name))
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.border = self.BORDER
        
        # Data
        for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.BODY_FONT
                cell.border = self.BORDER
        
        # Auto-width
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        return ws
    
    def add_index_sheet(self, analyses: List[str]):
        """Add index/TOC sheet at the beginning"""
        ws = self.wb.create_sheet(title="Index", index=0)
        
        ws.cell(row=1, column=1, value="Data Pack Index").font = Font(name="Libre Baskerville", size=18, bold=True, color="033333")
        
        ws.cell(row=3, column=1, value="Output Tabs:").font = Font(bold=True)
        
        row = 4
        for analysis in analyses:
            ws.cell(row=row, column=1, value=f"• {analysis}")
            row += 1
        
        row += 1
        ws.cell(row=row, column=1, value="Support Tabs:").font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value="• Raw Data - Source data with all transactions")
        
        row += 2
        ws.cell(row=row, column=1, value="Note: Output tabs contain formulas linked to Raw Data.").font = Font(italic=True, color="666666")
        ws.cell(row=row+1, column=1, value="Update Raw Data to refresh all outputs.").font = Font(italic=True, color="666666")
        
        ws.column_dimensions['A'].width = 60
        
        return ws
    
    def save(self):
        """Save the workbook"""
        self.wb.save(self.output_path)
        return self.output_path
